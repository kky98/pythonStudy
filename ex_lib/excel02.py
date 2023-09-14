import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
# pip install openpyxl
workbook = openpyxl.load_workbook('test.xlsx')
worksheet = workbook.active
worksheet.merge_cells('D1:E1') # 셀 병합
worksheet['A1'].font = Font(bold=True, size=24) # 폰트
yellow_fill = PatternFill(start_color='FFFF00'
                           , end_color='FFFF00'
                           , fill_type='solid')  #배경색
worksheet['A1'].fill = yellow_fill
worksheet.column_dimensions['A'].width = 100 #열 너비
worksheet.row_dimensions[1].height =100      #1행 높이
for row in worksheet.iter_rows(values_only=True):
    print(row)
workbook.close()
workbook.save('test2.xlsx')